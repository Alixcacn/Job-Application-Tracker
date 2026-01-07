from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, JobApplication
from datetime import datetime
import pandas as pd
from io import BytesIO
from flask import send_file
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this' # Change for production
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///job_tracker.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
@login_required
def index():
    applications = JobApplication.query.filter_by(user_id=current_user.id).order_by(JobApplication.date_applied.desc()).all()
    return render_template('index.html', user=current_user, applications=applications)

@app.route('/applications', methods=['POST'])
@login_required
def add_application():
    try:
        data = request.form
        new_job = JobApplication(
            user_id=current_user.id,
            company=data.get('company'),
            job_title=data.get('job_title'),
            location=data.get('location'),
            date_applied=datetime.strptime(data.get('date_applied'), '%Y-%m-%d').date() if data.get('date_applied') else datetime.utcnow().date(),
            source=data.get('source'),
            job_link=data.get('job_link'),
            resume_version=data.get('resume_version'),
            current_status=data.get('current_status', 'Applied'),
            salary_range=data.get('salary_range'),
            recruiter_info=data.get('recruiter_info'),
            referral_contact=data.get('referral_contact'),
            notes=data.get('notes'),
            timezone=data.get('timezone', 'EST'),
            priority=int(data.get('priority', 5))
        )
        
        if data.get('last_contacted_date'):
            new_job.last_contacted_date = datetime.strptime(data.get('last_contacted_date'), '%Y-%m-%d').date()

        if data.get('next_follow_up_date'):
            new_job.next_follow_up_date = datetime.strptime(data.get('next_follow_up_date'), '%Y-%m-%d').date()
            
        db.session.add(new_job)
        db.session.commit()
        
        # Return the new list item or the full list fragment
        applications = JobApplication.query.filter_by(user_id=current_user.id).order_by(JobApplication.date_applied.desc()).all()
        return render_template('partials/application_list.html', applications=applications)
    except Exception as e:
        return f"<div class='alert alert-error'>Error adding job: {str(e)}</div>", 400

@app.route('/applications/<int:id>/delete', methods=['DELETE'])
@login_required
def delete_application(id):
    try:
        application = JobApplication.query.get_or_404(id)
        if application.user_id != current_user.id:
            return "Unauthorized", 403
            
        db.session.delete(application)
        db.session.commit()
        
        # Return Updated List
        applications = JobApplication.query.filter_by(user_id=current_user.id).order_by(JobApplication.date_applied.desc()).all()
        return render_template('partials/application_list.html', applications=applications)
    except Exception as e:
        return f"<div class='alert alert-error'>Error deleting job: {str(e)}</div>", 400

@app.route('/applications/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_application(id):
    application = JobApplication.query.get_or_404(id)
    if application.user_id != current_user.id:
        return "Unauthorized", 403

    if request.method == 'POST':
        try:
            data = request.form
            application.company = data.get('company')
            application.job_title = data.get('job_title')
            application.location = data.get('location')
            if data.get('date_applied'):
                application.date_applied = datetime.strptime(data.get('date_applied'), '%Y-%m-%d').date()
            application.source = data.get('source')
            application.job_link = data.get('job_link')
            application.current_status = data.get('current_status')
            application.salary_range = data.get('salary_range')
            application.notes = data.get('notes')
            application.resume_version = data.get('resume_version')
            application.recruiter_info = data.get('recruiter_info')
            application.referral_contact = data.get('referral_contact')
            application.priority = int(data.get('priority', 5))

            if data.get('last_contacted_date'):
               application.last_contacted_date = datetime.strptime(data.get('last_contacted_date'), '%Y-%m-%d').date()
            else:
               application.last_contacted_date = None

            if data.get('next_follow_up_date'):
               application.next_follow_up_date = datetime.strptime(data.get('next_follow_up_date'), '%Y-%m-%d').date()
            else:
               application.next_follow_up_date = None
            
            db.session.commit()
            
            # Return Updated List
            applications = JobApplication.query.filter_by(user_id=current_user.id).order_by(JobApplication.date_applied.desc()).all()
            return render_template('partials/application_list.html', applications=applications)
        except Exception as e:
            return f"<div class='alert alert-error'>Error updating job: {str(e)}</div>", 400
    
    return render_template('partials/edit_job_form.html', app=application)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/export')
@login_required
def export_data():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        applications = JobApplication.query.filter_by(user_id=current_user.id).all()
        
        # 1. Define Columns
        columns = [
            'Company', 'Job Title', 'Location', 'Date Applied', 'Source', 
            'Job Link/JD', 'Resume Version', 'Current Status', 'Last Contacted', 
            'Recruiter/HM', 'Salary/TC Range', 'Referral/Contact', 'Interview Notes', 'Priority(1-10)'
        ]

        # Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        
        # Write Header
        ws.append(columns)

        # Write Data
        for app in applications:
            row = [
                app.company,
                app.job_title,
                app.location,
                app.date_applied.strftime('%Y-%m-%d') if app.date_applied else None,
                app.source,
                app.job_link,
                app.resume_version,
                app.current_status,
                app.last_contacted_date.strftime('%Y-%m-%d') if app.last_contacted_date else None,
                app.recruiter_info,
                app.salary_range,
                app.referral_contact,
                app.notes,
                app.priority
            ]
            ws.append(row)

        # --- Styling & Validation (User's Script) ---
        
        # Colors
        rainbow_colors = [
            "EBF5FB", "E9F7EF", "FEF9E7", "F5EEF8", 
            "FDEDEC", "FEF5E7", "E8F8F5", "F2F4F4"
        ]

        # Validations
        status_options = '"Applied,OA,Interview,Offer,Rejected,Ghosted"'
        source_options = '"LinkedIn,Indeed,Company Website,Referral,Handshake,Recruiter"'

        dv_status = DataValidation(type="list", formula1=status_options, allow_blank=True)
        ws.add_data_validation(dv_status)
        dv_status.add("H2:H500") 

        dv_source = DataValidation(type="list", formula1=source_options, allow_blank=True)
        ws.add_data_validation(dv_source)
        dv_source.add("E2:E500")

        # Header Style
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin', color="DDDDDD"), 
                             right=Side(style='thin', color="DDDDDD"), 
                             top=Side(style='thin', color="DDDDDD"), 
                             bottom=Side(style='thin', color="DDDDDD"))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Row Coloring
        # Determine max row to style (at least 500 as per request, or actual data length + buffer)
        max_row = max(500, len(applications) + 20)
        
        for row_idx in range(2, max_row + 1):
            color_index = (row_idx - 2) % len(rainbow_colors)
            current_fill = PatternFill(start_color=rainbow_colors[color_index], 
                                       end_color=rainbow_colors[color_index], 
                                       fill_type="solid")
            
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Only apply fill if validation allows or just overwrite? 
                # User script overwrites fill for all cells in range
                cell.fill = current_fill
                cell.border = thin_border

        # Formatting
        ws.freeze_panes = 'A2'
        column_widths = {'A': 22, 'B': 25, 'E': 18, 'H': 18, 'M': 50}
        for col, width in column_widths.items():
            if col in ws.column_dimensions:
                ws.column_dimensions[col].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, 
            download_name=f'job_applications_rainbow_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
         return f"<div class='alert alert-error'>Export failed: {str(e)}</div>", 500

@app.route('/import', methods=['POST'])
@login_required
def import_data():
    try:
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(url_for('index'))
            
        file = request.files['file']
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(url_for('index'))
            
        if file and file.filename.endswith('.xlsx'):
            df = pd.read_excel(file)
            
            # Map Excel Columns to DB Fields
            # 'Company', 'Job Title', 'Location', 'Date Applied', 'Source', ...
            count = 0
            for index, row in df.iterrows():
                # Skip empty rows (check required fields)
                if pd.isna(row.get('Company')) or pd.isna(row.get('Job Title')):
                    continue
                    
                # Create Application
                app = JobApplication(
                    user_id=current_user.id,
                    company=str(row.get('Company')),
                    job_title=str(row.get('Job Title')),
                    location=str(row.get('Location')) if not pd.isna(row.get('Location')) else None,
                    source=str(row.get('Source')) if not pd.isna(row.get('Source')) else None,
                    job_link=str(row.get('Job Link/JD')) if not pd.isna(row.get('Job Link/JD')) else None,
                    resume_version=str(row.get('Resume Version')) if not pd.isna(row.get('Resume Version')) else None,
                    current_status=str(row.get('Current Status')) if not pd.isna(row.get('Current Status')) else 'Applied',
                    recruiter_info=str(row.get('Recruiter/HM')) if not pd.isna(row.get('Recruiter/HM')) else None,
                    salary_range=str(row.get('Salary/TC Range')) if not pd.isna(row.get('Salary/TC Range')) else None,
                    referral_contact=str(row.get('Referral/Contact')) if not pd.isna(row.get('Referral/Contact')) else None,
                    notes=str(row.get('Interview Notes')) if not pd.isna(row.get('Interview Notes')) else None,
                    priority=int(row.get('Priority(1-10)')) if not pd.isna(row.get('Priority(1-10)')) and str(row.get('Priority(1-10)')).isdigit() else 5
                )

                # Date parsing is tricky, try basic Pandas to_datetime
                try:
                   if not pd.isna(row.get('Date Applied')):
                       app.date_applied = pd.to_datetime(row.get('Date Applied')).date()
                   if not pd.isna(row.get('Last Contacted')):
                       app.last_contacted_date = pd.to_datetime(row.get('Last Contacted')).date()
                except:
                    pass # Keep defaults or None if parse fails

                db.session.add(app)
                count += 1
            
            db.session.commit()
            flash(f'Successfully imported {count} applications.', 'success')
            return redirect(url_for('index'))
            
    except Exception as e:
        flash(f'Import failed: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/register', methods=['GET', 'POST']) # Temporary registration route
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return redirect(url_for('register'))
            
        hashed_password = generate_password_hash(password)
        new_user = User(username=username, password=hashed_password)
        db.session.add(new_user)
        db.session.commit()
        
        login_user(new_user)
        return redirect(url_for('index'))
    return render_template('register.html')

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)
